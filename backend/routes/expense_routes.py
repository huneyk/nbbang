"""사용자 스코프 경비/트립/설정/리포트 라우트.

모든 엔드포인트는 JWT 로그인이 필요하며, 현재 사용자의 active 트립을
컨텍스트로 동작한다. Google/한국수출입은행 API 키는 admin이 관리하므로
사용자 settings 응답에서는 제외된다.
"""
import os
import io
import uuid
import logging
import tempfile
from datetime import datetime
from flask import Blueprint, request, jsonify, make_response
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    SimpleDocTemplate, Image as RLImage, Table, TableStyle, Paragraph, PageBreak,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image as PILImage

from config import Config
from decorators import login_required, get_current_user_id
from services.ocr_service import analyze_receipt_with_gemini, calculate_krw_amount
from services import trip_repository, user_repository
from services.image_service import (
    get_image_info, fix_orientation_only, crop_receipt, downsize_for_storage,
)
from services.receipt_storage import (
    save_receipt as save_receipt_to_gridfs,
    get_receipt as get_receipt_from_gridfs,
)

logger = logging.getLogger(__name__)

expense_bp = Blueprint('expense', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
SENSITIVE_KEYS = ('google_api_key', 'koreaexim_api_key')


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _strip_sensitive(settings: dict) -> dict:
    return {k: v for k, v in (settings or {}).items() if k not in SENSITIVE_KEYS}


# ===== 영수증 업로드/조회 =====

@expense_bp.route('/api/upload-receipt', methods=['POST'])
@login_required
def upload_receipt():
    """영수증 이미지를 업로드하고 자동 crop → downsize → OCR 분석 → GridFS 저장."""
    if 'receipt' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400

    file = request.files['receipt']

    if file.filename == '':
        return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'}), 400

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '지원하지 않는 파일 형식입니다.'}), 400

    upload_folder = Config.UPLOAD_FOLDER
    os.makedirs(upload_folder, exist_ok=True)

    filename = secure_filename(file.filename)
    base_name = os.path.splitext(filename)[0]
    base_id = str(uuid.uuid4())

    temp_path = os.path.join(upload_folder, f"temp_{base_id}")
    oriented_path = os.path.join(upload_folder, f"oriented_{base_id}.jpg")
    cropped_path = os.path.join(upload_folder, f"cropped_{base_id}.jpg")
    storage_path = os.path.join(upload_folder, f"storage_{base_id}.jpg")

    temp_files = [temp_path, oriented_path, cropped_path, storage_path]
    file.save(temp_path)

    try:
        original_info = get_image_info(temp_path)
        if original_info:
            logger.info(
                f"원본 이미지: {original_info['width']}x{original_info['height']}, "
                f"{original_info['file_size_kb']}KB"
            )

        fix_orientation_only(temp_path, oriented_path)
        is_cropped = crop_receipt(oriented_path, cropped_path)
        logger.info(f"영수증 자동 crop: {'성공' if is_cropped else '윤곽 미감지 (원본 사용)'}")
        downsize_for_storage(cropped_path, storage_path)

        result = analyze_receipt_with_gemini(storage_path)

        storage_filename = f"{base_id}_{base_name}.jpg"
        receipt_id = save_receipt_to_gridfs(storage_path, storage_filename)

        if result['success']:
            result['data']['receipt_image'] = receipt_id

        return jsonify(result)

    except Exception as e:
        logger.error(f"영수증 처리 오류: {str(e)}")
        return jsonify({'success': False, 'error': f'영수증 처리에 실패했습니다: {str(e)}'}), 500

    finally:
        for p in temp_files:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except OSError:
                pass


@expense_bp.route('/api/receipts/<receipt_id>/image', methods=['GET'])
@login_required
def get_receipt_image(receipt_id: str):
    grid_out = get_receipt_from_gridfs(receipt_id)
    if grid_out is None:
        return jsonify({'success': False, 'error': '영수증 이미지를 찾을 수 없습니다.'}), 404

    response = make_response(grid_out.read())
    response.headers['Content-Type'] = grid_out.content_type or 'image/jpeg'
    response.headers['Content-Disposition'] = f'inline; filename="{grid_out.filename}"'
    response.headers['Cache-Control'] = 'private, max-age=31536000'
    return response


# ===== 경비 CRUD (현재 사용자 active 트립) =====

@expense_bp.route('/api/expenses', methods=['GET'])
@login_required
def get_expenses():
    user_id = get_current_user_id()
    expenses = trip_repository.list_expenses(user_id)
    return jsonify({'success': True, 'data': expenses})


@expense_bp.route('/api/expenses', methods=['POST'])
@login_required
def create_expense():
    user_id = get_current_user_id()
    data = request.json or {}

    required_fields = ['date', 'category', 'amount', 'currency', 'payment_method', 'payer']
    for field in required_fields:
        if field not in data:
            return jsonify({'success': False, 'error': f'{field} 필드가 필요합니다.'}), 400

    is_personal_expense = data.get('is_personal_expense', False)
    personal_expense_for = data.get('personal_expense_for') if is_personal_expense else None

    if is_personal_expense and not personal_expense_for:
        return jsonify({'success': False, 'error': '개인 지출 해당자를 선택해주세요.'}), 400

    settings = trip_repository.get_active_settings(user_id)
    user_fee_rate = user_repository.get_credit_card_fee_rate(user_id)
    krw_amount, exchange_rate = calculate_krw_amount(
        float(data['amount']), data['currency'], data['payment_method'], settings,
        credit_card_fee_rate=user_fee_rate,
    )

    expense = {
        'id': str(uuid.uuid4()),
        'date': data['date'],
        'category': data['category'],
        'amount': float(data['amount']),
        'currency': data['currency'],
        'payment_method': data['payment_method'],
        'krw_amount': krw_amount,
        'description': data.get('description', ''),
        'payer': data['payer'],
        'receipt_image': data.get('receipt_image'),
        'is_personal_expense': is_personal_expense,
        'personal_expense_for': personal_expense_for,
        'exchange_rate': exchange_rate,
        'created_at': datetime.utcnow().isoformat(),
    }

    saved = trip_repository.add_expense(user_id, expense)
    return jsonify({'success': True, 'data': saved}), 201


@expense_bp.route('/api/expenses/<expense_id>', methods=['DELETE'])
@login_required
def delete_expense(expense_id: str):
    user_id = get_current_user_id()
    if not trip_repository.delete_expense(user_id, expense_id):
        return jsonify({'success': False, 'error': '경비를 찾을 수 없습니다.'}), 404
    return jsonify({'success': True, 'message': '삭제되었습니다.'})


# ===== 요약/정산 =====

def _compute_summary(user_id: str) -> dict:
    settings = trip_repository.get_active_settings(user_id)
    expenses = trip_repository.list_expenses(user_id)

    participants = settings.get('participants', [])
    categories = settings.get('categories', [])

    shared_expenses = [e for e in expenses if not e.get('is_personal_expense', False)]
    personal_expenses = [e for e in expenses if e.get('is_personal_expense', False)]

    total_krw = sum(e.get('krw_amount', 0) for e in expenses)
    shared_total_krw = sum(e.get('krw_amount', 0) for e in shared_expenses)
    personal_total_krw = sum(e.get('krw_amount', 0) for e in personal_expenses)

    payer_totals = {p: 0 for p in participants}
    for exp in shared_expenses:
        payer = exp.get('payer', '')
        if payer in payer_totals:
            payer_totals[payer] += exp.get('krw_amount', 0)

    personal_expense_totals = {p: 0 for p in participants}
    personal_expense_details = {p: [] for p in participants}
    for exp in personal_expenses:
        person_for = exp.get('personal_expense_for', '')
        if person_for in personal_expense_totals:
            personal_expense_totals[person_for] += exp.get('krw_amount', 0)
            personal_expense_details[person_for].append({
                'date': exp.get('date', ''),
                'description': exp.get('description', ''),
                'amount': exp.get('krw_amount', 0),
                'payer': exp.get('payer', ''),
                'category': exp.get('category', ''),
            })

    paid_for_others = {p: 0 for p in participants}
    for exp in personal_expenses:
        payer = exp.get('payer', '')
        if payer in paid_for_others:
            paid_for_others[payer] += exp.get('krw_amount', 0)

    num_participants = len(participants)
    per_person = round(shared_total_krw / num_participants) if num_participants > 0 else 0

    settlements = {}
    for payer in participants:
        paid = payer_totals[payer]
        paid_for_other = paid_for_others[payer]
        personal_expense = personal_expense_totals[payer]
        shared_diff = per_person - paid
        final_diff = shared_diff + personal_expense - paid_for_other

        settlements[payer] = {
            'paid': round(paid),
            'paid_for_others': round(paid_for_other),
            'should_pay': per_person,
            'personal_expense': round(personal_expense),
            'personal_expense_details': personal_expense_details[payer],
            'shared_difference': round(shared_diff),
            'difference': round(final_diff),
            'status': '내야 할 금액' if final_diff > 0 else ('받을 금액' if final_diff < 0 else '정산 완료'),
        }

    category_totals = {c: 0 for c in categories}
    for exp in expenses:
        category = exp.get('category', '기타')
        if category in category_totals:
            category_totals[category] += exp.get('krw_amount', 0)

    return {
        'expenses': expenses,
        'shared_expenses': shared_expenses,
        'personal_expenses': personal_expenses,
        'settings': settings,
        'data': {
            'total_krw': round(total_krw),
            'shared_total_krw': round(shared_total_krw),
            'personal_total_krw': round(personal_total_krw),
            'per_person': per_person,
            'num_participants': num_participants,
            'payer_totals': {k: round(v) for k, v in payer_totals.items()},
            'settlements': settlements,
            'category_totals': {k: round(v) for k, v in category_totals.items()},
            'expense_count': len(expenses),
            'shared_expense_count': len(shared_expenses),
            'personal_expense_count': len(personal_expenses),
        },
    }


@expense_bp.route('/api/summary', methods=['GET'])
@login_required
def get_summary():
    user_id = get_current_user_id()
    return jsonify({'success': True, 'data': _compute_summary(user_id)['data']})


# ===== 설정/통화/환율 =====

@expense_bp.route('/api/config', methods=['GET'])
@login_required
def get_config():
    user_id = get_current_user_id()
    settings = trip_repository.get_active_settings(user_id)
    user_fee_rate = user_repository.get_credit_card_fee_rate(user_id)
    return jsonify({
        'success': True,
        'data': {
            'trip_title': settings.get('trip_title', '여행 경비 정산'),
            'participants': settings.get('participants', []),
            'categories': settings.get('categories', []),
            'currencies': settings.get('currencies', []),
            'exchange_rates': settings.get('exchange_rates', {}),
            'credit_card_fee_rate': user_fee_rate,
            'exchange_rate_info': settings.get('exchange_rate_info', {}),
        },
    })


@expense_bp.route('/api/settings', methods=['GET'])
@login_required
def get_settings():
    user_id = get_current_user_id()
    settings = _strip_sensitive(trip_repository.get_active_settings(user_id))
    # 수수료율은 트립이 아닌 사용자 프로필 값을 최종 소스로 사용한다.
    settings['credit_card_fee_rate'] = user_repository.get_credit_card_fee_rate(user_id)
    return jsonify({'success': True, 'data': settings})


@expense_bp.route('/api/settings', methods=['PUT'])
@login_required
def update_settings():
    user_id = get_current_user_id()
    data = request.json or {}
    settings = trip_repository.get_active_settings(user_id)

    if 'trip_title' in data:
        settings['trip_title'] = data['trip_title']
    if 'participants' in data:
        settings['participants'] = [p.strip() for p in data['participants'] if p.strip()]
    if 'categories' in data:
        settings['categories'] = [c.strip() for c in data['categories'] if c.strip()]

    # 신용카드 수수료율은 사용자 프로필에 저장한다(트립 공용 값이 아님).
    if 'credit_card_fee_rate' in data:
        user_repository.set_credit_card_fee_rate(
            user_id, data['credit_card_fee_rate'],
        )

    if 'currencies' in data:
        settings['currencies'] = data['currencies']
        settings['exchange_rates'] = {c['code']: float(c['rate']) for c in data['currencies']}
    elif 'exchange_rates' in data:
        settings['exchange_rates'] = {k: float(v) for k, v in data['exchange_rates'].items()}

    settings = _strip_sensitive(settings)
    trip_repository.save_active_settings(user_id, settings)
    settings['credit_card_fee_rate'] = user_repository.get_credit_card_fee_rate(user_id)
    return jsonify({'success': True, 'data': settings})


@expense_bp.route('/api/exchange-rates', methods=['PUT'])
@login_required
def update_exchange_rates():
    user_id = get_current_user_id()
    data = request.json or {}
    settings = trip_repository.get_active_settings(user_id)

    exchange_rates = settings.get('exchange_rates', {})
    currencies = settings.get('currencies', [])

    for currency_code, rate in data.items():
        exchange_rates[currency_code] = float(rate)
        for curr in currencies:
            if curr['code'] == currency_code:
                curr['rate'] = float(rate)
                break

    settings['exchange_rates'] = exchange_rates
    settings['currencies'] = currencies
    trip_repository.save_active_settings(user_id, settings)
    return jsonify({'success': True, 'data': exchange_rates})


@expense_bp.route('/api/exchange-rates/fetch', methods=['POST'])
@login_required
def fetch_latest_exchange_rates():
    from services.exchange_rate_service import fetch_exchange_rates, apply_fetched_rates

    user_id = get_current_user_id()
    settings = trip_repository.get_active_settings(user_id)
    result = fetch_exchange_rates(settings)

    if not result.get('rates'):
        return jsonify({
            'success': False,
            'error': '환율 조회에 실패했습니다. 잠시 후 다시 시도해주세요.',
        }), 502

    updated = apply_fetched_rates(settings, result)
    trip_repository.save_active_settings(user_id, updated)

    return jsonify({
        'success': True,
        'data': {
            'rates': result['rates'],
            'source': result['source'],
            'updated_at': result['updated_at'],
            'rate_type': result['rate_type'],
            'currencies': updated.get('currencies', []),
        },
    })


@expense_bp.route('/api/exchange-rates/info', methods=['GET'])
@login_required
def get_exchange_rate_info():
    user_id = get_current_user_id()
    settings = trip_repository.get_active_settings(user_id)
    return jsonify({'success': True, 'data': settings.get('exchange_rate_info', {})})


# ===== 통화 관리 =====

@expense_bp.route('/api/currencies', methods=['GET'])
@login_required
def get_currencies():
    from services.exchange_rate_service import CURRENCY_INFO, DEFAULT_FETCH_CURRENCIES

    user_id = get_current_user_id()
    settings = trip_repository.get_active_settings(user_id)
    currencies = settings.get('currencies') or list(trip_repository.DEFAULT_CURRENCIES)

    existing_codes = {c['code'] for c in currencies}
    default_rates = {c['code']: c['rate'] for c in trip_repository.DEFAULT_CURRENCIES}
    added = False

    for code in DEFAULT_FETCH_CURRENCIES:
        if code not in existing_codes:
            info = CURRENCY_INFO.get(code, {'name': code, 'flag': '🏳️'})
            currencies.append({
                'code': code,
                'name': info['name'],
                'flag': info['flag'],
                'rate': default_rates.get(code, 1.0),
                'is_base': False,
            })
            added = True

    if added:
        settings['currencies'] = currencies
        trip_repository.save_active_settings(user_id, settings)

    return jsonify({'success': True, 'data': currencies})


@expense_bp.route('/api/currencies', methods=['POST'])
@login_required
def add_currency():
    user_id = get_current_user_id()
    data = request.json or {}

    if not data.get('code') or not data.get('name'):
        return jsonify({'success': False, 'error': '통화 코드와 이름은 필수입니다.'}), 400

    code = data['code'].upper().strip()
    settings = trip_repository.get_active_settings(user_id)
    currencies = settings.get('currencies', [])
    exchange_rates = settings.get('exchange_rates', {'KRW': 1.0})

    if any(c['code'] == code for c in currencies):
        return jsonify({'success': False, 'error': f'통화 코드 {code}가 이미 존재합니다.'}), 400

    new_currency = {
        'code': code,
        'name': data['name'].strip(),
        'flag': data.get('flag', '🏳️').strip(),
        'rate': float(data.get('rate', 1.0)),
        'is_base': False,
    }
    currencies.append(new_currency)
    exchange_rates[code] = new_currency['rate']

    settings['currencies'] = currencies
    settings['exchange_rates'] = exchange_rates
    trip_repository.save_active_settings(user_id, settings)
    return jsonify({'success': True, 'data': new_currency}), 201


@expense_bp.route('/api/currencies/<currency_code>', methods=['PUT'])
@login_required
def update_currency(currency_code: str):
    user_id = get_current_user_id()
    data = request.json or {}
    settings = trip_repository.get_active_settings(user_id)
    currencies = settings.get('currencies', [])
    exchange_rates = settings.get('exchange_rates', {})

    currency = next((c for c in currencies if c['code'] == currency_code.upper()), None)
    if not currency:
        return jsonify({'success': False, 'error': '통화를 찾을 수 없습니다.'}), 404

    if currency.get('is_base') and 'rate' in data and float(data['rate']) != 1.0:
        return jsonify({'success': False, 'error': '기준 통화의 환율은 변경할 수 없습니다.'}), 400

    if 'name' in data:
        currency['name'] = data['name'].strip()
    if 'flag' in data:
        currency['flag'] = data['flag'].strip()
    if 'rate' in data and not currency.get('is_base'):
        currency['rate'] = float(data['rate'])
        exchange_rates[currency_code.upper()] = float(data['rate'])

    settings['currencies'] = currencies
    settings['exchange_rates'] = exchange_rates
    trip_repository.save_active_settings(user_id, settings)
    return jsonify({'success': True, 'data': currency})


@expense_bp.route('/api/currencies/<currency_code>', methods=['DELETE'])
@login_required
def delete_currency(currency_code: str):
    user_id = get_current_user_id()
    settings = trip_repository.get_active_settings(user_id)
    currencies = settings.get('currencies', [])
    exchange_rates = settings.get('exchange_rates', {})

    currency = next((c for c in currencies if c['code'] == currency_code.upper()), None)
    if not currency:
        return jsonify({'success': False, 'error': '통화를 찾을 수 없습니다.'}), 404
    if currency.get('is_base'):
        return jsonify({'success': False, 'error': '기준 통화는 삭제할 수 없습니다.'}), 400

    currencies = [c for c in currencies if c['code'] != currency_code.upper()]
    exchange_rates.pop(currency_code.upper(), None)

    settings['currencies'] = currencies
    settings['exchange_rates'] = exchange_rates
    trip_repository.save_active_settings(user_id, settings)
    return jsonify({'success': True, 'message': f'{currency_code} 통화가 삭제되었습니다.'})


# ===== 트립 관리 =====

@expense_bp.route('/api/trips', methods=['GET'])
@login_required
def get_trips():
    user_id = get_current_user_id()
    return jsonify({'success': True, 'data': trip_repository.list_trips(user_id)})


@expense_bp.route('/api/trips/new', methods=['POST'])
@login_required
def create_trip():
    """새 트립 생성. 현재 활성 트립은 자동으로 비활성화된다."""
    user_id = get_current_user_id()
    data = request.json or {}

    # 신용카드 수수료율은 사용자 프로필에 저장되므로 트립 단위로는 받지 않는다.
    # (구 버전 호환을 위해 payload에 포함되어도 무시한다.)
    new_trip = trip_repository.create_trip(
        user_id=user_id,
        trip_title=data.get('trip_title', '새 여행'),
        participants=data.get('participants') or None,
        categories=data.get('categories') or None,
        credit_card_fee_rate=user_repository.get_credit_card_fee_rate(user_id),
        make_active=True,
    )
    new_settings = _strip_sensitive(new_trip['settings'])
    new_settings['credit_card_fee_rate'] = user_repository.get_credit_card_fee_rate(user_id)
    return jsonify({
        'success': True,
        'data': {
            'trip_id': new_trip['trip_id'],
            'new_settings': new_settings,
        },
    }), 201


@expense_bp.route('/api/trips/<trip_id>', methods=['GET'])
@login_required
def get_trip(trip_id: str):
    """선택한 트립을 active로 전환하고 settings + expenses 반환."""
    user_id = get_current_user_id()
    trip = trip_repository.set_active_trip(user_id, trip_id)
    if not trip:
        return jsonify({'success': False, 'error': '여행을 찾을 수 없습니다.'}), 404
    settings = _strip_sensitive(trip.get('settings', {}))
    settings['credit_card_fee_rate'] = user_repository.get_credit_card_fee_rate(user_id)
    return jsonify({
        'success': True,
        'data': {
            'settings': settings,
            'expenses': trip.get('expenses', []),
        },
    })


@expense_bp.route('/api/trips/<trip_id>', methods=['DELETE'])
@login_required
def remove_trip(trip_id: str):
    user_id = get_current_user_id()
    if not trip_repository.delete_trip(user_id, trip_id):
        return jsonify({'success': False, 'error': '여행을 찾을 수 없습니다.'}), 404
    return jsonify({'success': True, 'message': '여행이 삭제되었습니다.'})


# ===== Excel/PDF 리포트 =====

@expense_bp.route('/api/report/download', methods=['GET'])
@login_required
def download_report():
    user_id = get_current_user_id()
    summary = _compute_summary(user_id)
    expenses = sorted(summary['expenses'], key=lambda e: e.get('date', ''))
    settings = summary['settings']
    data = summary['data']

    participants = settings.get('participants', [])
    categories = settings.get('categories', [])
    trip_title = settings.get('trip_title', '여행 경비 정산')

    wb = Workbook()
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    personal_fill = PatternFill(start_color='fff0f0', end_color='fff0f0', fill_type='solid')
    currency_font = Font(name='Consolas', size=11)
    border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    ws_expenses = wb.active
    ws_expenses.title = "경비 내역"

    ws_expenses.merge_cells('A1:J1')
    ws_expenses['A1'] = f'🌏 {trip_title} - 경비 내역서'
    ws_expenses['A1'].font = Font(bold=True, size=18, color='e94560')
    ws_expenses['A1'].alignment = center_align
    ws_expenses.row_dimensions[1].height = 35

    ws_expenses.merge_cells('A2:J2')
    ws_expenses['A2'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_expenses['A2'].alignment = center_align
    ws_expenses['A2'].font = Font(size=10, color='666666')

    headers = ['날짜', '지출 항목', '금액', '통화', '결제수단', '적용 환율',
               '원화 환산액', '세부 내역', '지불한 사람', '지출 유형']
    for col, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row, exp in enumerate(expenses, 5):
        is_personal = exp.get('is_personal_expense', False)
        ws_expenses.cell(row=row, column=1, value=exp.get('date', '')).border = border
        ws_expenses.cell(row=row, column=2, value=exp.get('category', '')).border = border

        amount_cell = ws_expenses.cell(row=row, column=3, value=exp.get('amount', 0))
        amount_cell.font = currency_font
        amount_cell.alignment = right_align
        amount_cell.border = border
        amount_cell.number_format = '#,##0'

        ws_expenses.cell(row=row, column=4, value=exp.get('currency', '')).border = border
        ws_expenses.cell(row=row, column=4).alignment = center_align

        ws_expenses.cell(row=row, column=5, value=exp.get('payment_method', '')).border = border
        ws_expenses.cell(row=row, column=5).alignment = center_align

        exchange_rate_val = exp.get('exchange_rate')
        if exchange_rate_val is None and exp.get('amount') and exp.get('krw_amount'):
            exchange_rate_val = round(exp['krw_amount'] / exp['amount'], 2) if exp['amount'] != 0 else None
        rate_cell = ws_expenses.cell(row=row, column=6, value=exchange_rate_val or '')
        rate_cell.font = Font(name='Consolas', size=11)
        rate_cell.alignment = right_align
        rate_cell.border = border
        if exchange_rate_val:
            rate_cell.number_format = '#,##0.00'

        krw_cell = ws_expenses.cell(row=row, column=7, value=exp.get('krw_amount', 0))
        krw_cell.font = Font(name='Consolas', size=11, color='ffc300')
        krw_cell.alignment = right_align
        krw_cell.border = border
        krw_cell.number_format = '₩#,##0'

        ws_expenses.cell(row=row, column=8, value=exp.get('description', '')).border = border
        ws_expenses.cell(row=row, column=9, value=exp.get('payer', '')).border = border

        expense_type = f"개인 ({exp.get('personal_expense_for', '')})" if is_personal else "공동"
        type_cell = ws_expenses.cell(row=row, column=10, value=expense_type)
        type_cell.border = border
        type_cell.alignment = center_align
        if is_personal:
            type_cell.font = Font(color='e94560')
            for c in range(1, 11):
                ws_expenses.cell(row=row, column=c).fill = personal_fill

    column_widths = [12, 12, 15, 8, 12, 14, 18, 25, 12, 14]
    for i, width in enumerate(column_widths, 1):
        ws_expenses.column_dimensions[get_column_letter(i)].width = width

    ws_summary = wb.create_sheet(title="정산 요약")
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = '💰 경비 정산 요약'
    ws_summary['A1'].font = Font(bold=True, size=18, color='e94560')
    ws_summary['A1'].alignment = center_align

    ws_summary['A3'] = '총 경비'
    ws_summary['A3'].font = Font(bold=True, size=12)
    ws_summary['B3'] = data['total_krw']
    ws_summary['B3'].font = Font(name='Consolas', size=14, bold=True, color='ffc300')
    ws_summary['B3'].number_format = '₩#,##0'
    ws_summary['C3'] = f"({data['expense_count']}건)"

    ws_summary['A4'] = '공동 경비'
    ws_summary['A4'].font = Font(bold=True, size=12)
    ws_summary['B4'] = data['shared_total_krw']
    ws_summary['B4'].font = Font(name='Consolas', size=12, color='00d9a5')
    ws_summary['B4'].number_format = '₩#,##0'
    ws_summary['C4'] = f"({data['shared_expense_count']}건)"

    ws_summary['A5'] = '개인 지출'
    ws_summary['A5'].font = Font(bold=True, size=12)
    ws_summary['B5'] = data['personal_total_krw']
    ws_summary['B5'].font = Font(name='Consolas', size=12, color='e94560')
    ws_summary['B5'].number_format = '₩#,##0'
    ws_summary['C5'] = f"({data['personal_expense_count']}건)"

    ws_summary['A6'] = '1인당 분담액 (공동 경비)'
    ws_summary['A6'].font = Font(bold=True, size=12)
    ws_summary['B6'] = data['per_person']
    ws_summary['B6'].font = Font(name='Consolas', size=14, bold=True, color='00d9a5')
    ws_summary['B6'].number_format = '₩#,##0'
    ws_summary['C6'] = f"({data['num_participants']}명 기준)"

    ws_summary['A8'] = '💸 정산 내역'
    ws_summary['A8'].font = Font(bold=True, size=14)

    settlement_headers = ['이름', '공동 경비 지불', '타인 개인지출 대납', '분담액',
                          '개인 지출', '최종 정산액', '상태']
    for col, header in enumerate(settlement_headers, 1):
        cell = ws_summary.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    row = 10
    for name, sd in data['settlements'].items():
        ws_summary.cell(row=row, column=1, value=name).border = border
        for col_idx, key in enumerate(
            ['paid', 'paid_for_others', 'should_pay', 'personal_expense', 'difference'], 2,
        ):
            cell = ws_summary.cell(row=row, column=col_idx, value=sd[key])
            cell.number_format = '₩#,##0'
            cell.alignment = right_align
            cell.border = border
        diff_cell = ws_summary.cell(row=row, column=6)
        if sd['difference'] > 0:
            diff_cell.font = Font(color='e94560', bold=True)
        elif sd['difference'] < 0:
            diff_cell.font = Font(color='00d9a5', bold=True)
        status_cell = ws_summary.cell(row=row, column=7, value=sd['status'])
        status_cell.alignment = center_align
        status_cell.border = border
        row += 1

    cat_start_row = row + 2
    ws_summary.cell(row=cat_start_row, column=1, value='📊 카테고리별 지출').font = Font(bold=True, size=14)

    cat_headers = ['카테고리', '금액', '비율']
    for col, header in enumerate(cat_headers, 1):
        cell = ws_summary.cell(row=cat_start_row + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    sorted_categories = sorted(data['category_totals'].items(), key=lambda x: x[1], reverse=True)
    row = cat_start_row + 2
    total_krw = data['total_krw']
    for category, amount in sorted_categories:
        if amount > 0:
            ws_summary.cell(row=row, column=1, value=category).border = border
            amount_cell = ws_summary.cell(row=row, column=2, value=amount)
            amount_cell.number_format = '₩#,##0'
            amount_cell.alignment = right_align
            amount_cell.border = border
            pct = (amount / total_krw) if total_krw > 0 else 0
            pct_cell = ws_summary.cell(row=row, column=3, value=pct)
            pct_cell.number_format = '0.0%'
            pct_cell.alignment = center_align
            pct_cell.border = border
            row += 1

    for col, w in zip('ABCDEFG', [20, 18, 18, 15, 15, 15, 12]):
        ws_summary.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"expense_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


_korean_font_registered = None


def _register_korean_font():
    """한글 폰트를 등록합니다."""
    global _korean_font_registered
    if _korean_font_registered is not None:
        return _korean_font_registered

    bundled_font = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'fonts', 'NotoSansKR-Regular.ttf',
    )
    candidates = [
        bundled_font,
        '/Library/Fonts/Arial Unicode.ttf',
        '/System/Library/Fonts/Supplemental/AppleGothic.ttf',
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        'C:/Windows/Fonts/malgun.ttf',
    ]
    for path in candidates:
        if not os.path.exists(path):
            continue
        try:
            pdfmetrics.registerFont(TTFont('Korean', path))
            logger.info(f"한글 폰트 등록 성공: {path}")
            _korean_font_registered = 'Korean'
            return 'Korean'
        except Exception as e:
            logger.warning(f"폰트 로드 실패 ({path}): {e}")

    logger.error("한글 폰트를 찾을 수 없습니다. Helvetica를 사용합니다.")
    _korean_font_registered = 'Helvetica'
    return 'Helvetica'


@expense_bp.route('/api/report/download-receipts', methods=['GET'])
@login_required
def download_receipt_report():
    user_id = get_current_user_id()
    expenses = sorted(trip_repository.list_expenses(user_id), key=lambda e: e.get('date', ''))
    settings = trip_repository.get_active_settings(user_id)
    trip_title = settings.get('trip_title', '여행 경비 정산')

    if not expenses:
        return jsonify({'success': False, 'error': '경비 내역이 없습니다.'}), 400

    korean_font = _register_korean_font()
    receipt_count = sum(1 for e in expenses if e.get('receipt_image'))

    page_w, page_h = A4
    margin = 10 * mm
    usable_w = page_w - 2 * margin
    usable_h = page_h - 2 * margin

    grid_cols = 2
    rows_per_page = 5
    col_gap = 4 * mm
    row_gap = 2 * mm

    cell_w = (usable_w - col_gap * (grid_cols - 1)) / grid_cols
    title_h = 16 * mm
    cell_h = (usable_h - title_h - row_gap * (rows_per_page - 1)) / rows_per_page

    label_col_w = cell_w * 0.48
    img_col_w = cell_w * 0.52 - 2 * mm

    field_label_style = ParagraphStyle(
        'FieldLabel', fontName=korean_font, fontSize=8, leading=10,
        textColor=HexColor('#2d5016'),
    )
    field_value_style = ParagraphStyle(
        'FieldValue', fontName=korean_font, fontSize=7.5, leading=9.5,
        textColor=HexColor('#333333'),
    )
    no_receipt_style = ParagraphStyle(
        'NoReceipt', fontName=korean_font, fontSize=9, leading=12,
        textColor=HexColor('#aaaaaa'), alignment=1,
    )
    title_style = ParagraphStyle(
        'ReceiptTitle', fontName=korean_font, fontSize=13, leading=17,
        textColor=HexColor('#1a1a2e'),
    )
    subtitle_style = ParagraphStyle(
        'ReceiptSubtitle', fontName=korean_font, fontSize=8, leading=11,
        textColor=HexColor('#888888'), spaceAfter=2 * mm,
    )

    field_labels = ['날짜', '지출 항목', '금액', '결제수단',
                    '적용 환율', '원화 환산액', '세부 내역', '지불한 사람']

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin,
    )

    elements = []
    temp_files = []

    try:
        cells = []
        for exp in expenses:
            receipt_id = exp.get('receipt_image')
            img_path = None

            if receipt_id:
                grid_out = get_receipt_from_gridfs(receipt_id)
                if grid_out is not None:
                    tmp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    tmp.write(grid_out.read())
                    tmp.flush()
                    tmp.close()
                    temp_files.append(tmp.name)
                    img_path = tmp.name

            amount_val = exp.get('amount', 0)
            currency = exp.get('currency', 'KRW')
            krw_amount = exp.get('krw_amount', 0)

            exchange_rate_val = exp.get('exchange_rate')
            if exchange_rate_val is None and amount_val and krw_amount and amount_val != 0:
                exchange_rate_val = round(krw_amount / amount_val, 2)
            exchange_str = f"{exchange_rate_val:,.2f}" if exchange_rate_val else ''

            description = exp.get('description', '')
            desc_short = description[:15] + ('…' if len(description) > 15 else '') if description else ''

            field_values = [
                exp.get('date', ''), exp.get('category', ''),
                f"{amount_val:,.0f} {currency}", exp.get('payment_method', ''),
                exchange_str, f"₩{krw_amount:,.0f}",
                desc_short, exp.get('payer', ''),
            ]

            label_rows = [
                [Paragraph(lbl, field_label_style), Paragraph(str(val), field_value_style)]
                for lbl, val in zip(field_labels, field_values)
            ]

            inner_h = cell_h - 2 * mm
            row_h = inner_h / len(field_labels)
            label_name_w = label_col_w * 0.48
            label_val_w = label_col_w * 0.52

            info_table = Table(
                label_rows,
                colWidths=[label_name_w, label_val_w],
                rowHeights=[row_h] * len(field_labels),
            )
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 1 * mm),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0.5 * mm),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))

            if img_path:
                try:
                    pil_img = PILImage.open(img_path)
                    orig_w, orig_h = pil_img.size
                    pil_img.close()
                except Exception:
                    orig_w, orig_h = 300, 400

                scale = min(img_col_w / orig_w, inner_h / orig_h)
                img_element = RLImage(img_path, width=orig_w * scale, height=orig_h * scale)
            else:
                placeholder = Table(
                    [[Paragraph('영수증 없음', no_receipt_style)]],
                    colWidths=[img_col_w], rowHeights=[inner_h],
                )
                placeholder.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('BACKGROUND', (0, 0), (-1, -1), HexColor('#f5f5f5')),
                ]))
                img_element = placeholder

            cell_content = Table(
                [[info_table, img_element]],
                colWidths=[label_col_w, img_col_w],
                rowHeights=[inner_h],
            )
            cell_content.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('BOX', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
            ]))
            cells.append(cell_content)

        items_per_page = grid_cols * rows_per_page
        total_pages = (len(cells) + items_per_page - 1) // items_per_page

        for page_idx in range(total_pages):
            if page_idx > 0:
                elements.append(PageBreak())

            elements.append(Paragraph(f'{trip_title} - 영수증 첨부', title_style))
            elements.append(Paragraph(
                f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                f"총 {len(expenses)}건 (영수증 {receipt_count}건)  |  "
                f"페이지 {page_idx + 1}/{total_pages}",
                subtitle_style,
            ))

            start = page_idx * items_per_page
            page_cells = cells[start:start + items_per_page]

            page_grid_rows = []
            for r in range(rows_per_page):
                row_cells = []
                for c in range(grid_cols):
                    idx = r * grid_cols + c
                    row_cells.append(page_cells[idx] if idx < len(page_cells) else '')
                page_grid_rows.append(row_cells)

            grid_table = Table(
                page_grid_rows,
                colWidths=[cell_w] * grid_cols,
                rowHeights=[cell_h] * rows_per_page,
                hAlign='CENTER',
            )
            grid_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), col_gap / 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), col_gap / 2),
                ('TOPPADDING', (0, 0), (-1, -1), row_gap / 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), row_gap / 2),
            ]))
            elements.append(grid_table)

        doc.build(elements)

    finally:
        for fp in temp_files:
            try:
                os.unlink(fp)
            except OSError:
                pass

    output.seek(0)
    filename = f"receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response
